from tkinter import *
from tkinter import messagebox
from tkinter import font

root=Tk()

def scoreinfo():
    messagebox.showinfo("","Score 1-3: You need to be improving on your health, try out some new activities!                                                                     Score 4-6: You're doing alright but always room for improvement                                                                                  Score 7-10: AMAZING! keep it up")
                        
import csv
def verify():
    found = False 
    file_r = open("usernames.csv" , "r")
    readCSV = csv.reader(file_r)
    
    myusername= entry1.get()
    mypassword= entry2.get()
    found="no"

    for row in readCSV:
        username = row[0]
        password = row[1]

        if username == myusername:
            found="yes"
            if password == mypassword:
                messagebox.showinfo("","Welcome")
                found = True

                #encryption
                import random
                plaintext = mypassword
                plaintext = plaintext.lower()
                ciphertext = ""

                for c in plaintext:
                    if c in "abcdefghijklmnopqrstuvwxyz":
                        num = ord(c)
                        randomNumber = random.randint(1,26)
                        num = num + randomNumber
                        if num > ord("z"):
                            #will subtract 26 because it wrap when over z
                            num = num-randomNumber
                        ciphertext = ciphertext+chr(num)
                        encryptedPassword = ciphertext
                    else:
                        ciphertext = ciphertext + c
                        encryptedPassword = ciphertext

                activitiesLabel = Label(root, text="Hey, what have you been up to today?").grid(row=17, column=1)

                sleepingVar = IntVar()
                sportsVar = IntVar()
                friendsVar = IntVar()
                exerciseVar = IntVar()
                relaxVar = IntVar()
                eatingVar = IntVar()
                shoppingVar = IntVar()
                readingVar = IntVar()
                studyVar = IntVar()

                
                def activitiesModule(average):
                    if sleepingVar == 1:
                        value = 4
                        totalValue = totalValue + value
                        numBoxes = numBoxes + 1
                    if sportsVar == 1:
                        value = 8
                        totalValue = totalValue + value
                        numBoxes = numBoxes + 1
                    if friendsVar == 1:
                        value = 5
                        totalValue = totalValue + value
                        numBoxes = numBoxes + 1
                    if exerciseVar == 1:
                        value = 9
                        totalValue = totalValue + value
                        numBoxes = numBoxes + 1
                    if relaxVar == 1:
                        value = 5
                        totalValue = totalValue + value
                        numBoxes = numBoxes + 1
                    if eatingVar == 1:
                        value = 7
                        totalValue = totalValue + value
                        numBoxes = numBoxes + 1
                    if shoppingVar == 1:
                        value = 4
                        totalValue = totalValue + value
                        numBoxes = numBoxes + 1
                    if readingVar == 1:
                        value = 7
                        totalValue = totalValue + value
                        numBoxes = numBoxes + 1
                    if studyVar == 1:
                        value = 7
                        totalValue = totalValue + value
                        numBoxes = numBoxes + 1
                        
                    return average
                    average = totalValue/numBoxes
                    scoreLabel = Label(root, text="Your average score is: " + average).grid(row=32, column=0)


                
                sleeping= Checkbutton(root, text="Sleeping",variable=sleepingVar).grid(row=18, column=0)
                sports= Checkbutton(root, text="Sports",variable=sportsVar).grid(row=19, column=0)
                friends= Checkbutton(root,text="Friends",variable=friendsVar).grid(row=20, column=0)
                exercise= Checkbutton(root,text="Exercising",variable=exerciseVar).grid(row=21, column=0)
                relax= Checkbutton(root,text="Relaxing",variable=relaxVar).grid(row=22, column=0)
                eating= Checkbutton(root,text="Eating healthily",variable=eatingVar).grid(row=24, column=0)
                shopping= Checkbutton(root,text="Shopping",variable=shoppingVar).grid(row=25, column=0)
                reading= Checkbutton(root,text="Reading",variable=readingVar).grid(row=26, column=0)
                study= Checkbutton(root,text="Studying",variable=studyVar).grid(row=27, column=0)
                Button(root,text="Submit",command=activitiesModule).grid(row=28, column=0)
                Button(root, text="Click here for score analysis", command=scoreinfo).grid(row=31, column=0)
                notesLabel = Label(root, text="Here are your post it notes").grid(row=0, column=50)
                Button(root, text="DIARY MODE", command=diaryMode).grid(row=7, column=50)
                notesSection()
                notesSection2()
                notesSection3()
                
            else:
                messagebox.showinfo("","Incorect username or password")

    if found=="no":
        messagebox.showinfo("","NOT REGISTERED")
    file_r.close()

def welcome():
    messagebox.showinfo("Post it Diary","This is an app which will allow you to take note of your daily life by recording and saving memories in which you'll never forget something again                                                        created by Jamie-Lee Gordon")

Button(root, text="Post it Diary",padx=60, pady=8, fg="white",bg="blue",font=("Calabri", 21), command=welcome).grid(row=0, column=1)


usernameLabel=Label(root, text="Username")
passwordLabel=Label(root, text="Password")
entry1=Entry(root)
entry2=Entry(root)
    
usernameLabel.grid(row=1, sticky=E)
passwordLabel.grid(row=2, sticky=E)

entry1.grid(row=1, column=1)
entry2.grid(row=2, column=1)

Button(root, text="login",padx=40,pady=2, command=verify, fg="black", bg="grey").grid(row=3, column=1)

check= Checkbutton(root, text="Remember me")
check.grid(columnspan=2)

def registerButton():
    from openpyxl import load_workbook
    filename = "usernamesAmmend.pyxl"
    wb = load_workbook(filename)
    ws = workbook.sheets[0]
    ws_tables = []

    #ammending the database
    ws["A6"] = entryUsernameLabel.get()
    ws["F6"] = entryUsernameLabel.get()
    ws["B6"] = entryPasswordLabel.get()
    ws["C6"] = entryFirstNameLabel.get()
    ws["G6"] = entryFirstNameLabel.get()
    ws["D6"] = entryLastNameLabel.get()
    ws["H6"] = entryLastNameLabel.get()
    ws["I6"] = entrydobLabel.get()
    ws["J6"] = entryPhoneLabel.get()
    ws["K6"] = entryEmailLabel.get()

    wb.save(filename)
    
def newUserButton():
    newUsernameLabel=Label(root, text="New username")
    newPasswordLabel=Label(root, text="New Password")
    reEnterPasswordLabel=Label(root, text="Re-enter password")
    firstNameLabel=Label(root, text="Enter first name")
    lastNameLabel=Label(root, text="Enter last name")
    dobLabel=Label(root, text="Enter date of birth")
    phoneLabel=Label(root, text="Enter phone number")
    emailLabel=Label(root, text="Enter Email")

    newUsernameLabel.grid(row=1, column=8)
    newPasswordLabel.grid(row=2, column=8)
    reEnterPasswordLabel.grid(row=3, column=8)
    firstNameLabel.grid(row=4, column=8)
    lastNameLabel.grid(row=5, column=8)
    dobLabel.grid(row=6, column=8)
    phoneLabel.grid(row=7, column=8)
    emailLabel.grid(row=8, column=8)

    entryUsernameLabel=Entry(root)
    entryPasswordLabel=Entry(root)
    entryReenterPasswordLabel=Entry(root)
    entryFirstNameLabel=Entry(root)
    entryLastNameLabel=Entry(root)
    entrydobLabel=Entry(root)
    entryPhoneLabel=Entry(root)
    entryEmailLabel=Entry(root)

    entryUsernameLabel.grid(row=1, column=9)
    entryPasswordLabel.grid(row=2, column=9)
    entryReenterPasswordLabel.grid(row=3, column=9)
    entryFirstNameLabel.grid(row=4, column=9)
    entryLastNameLabel.grid(row=5, column=9)
    entrydobLabel.grid(row=6, column=9)
    entryPhoneLabel.grid(row=7, column=9)
    entryEmailLabel.grid(row=8, column=9)

    Button(root, text="Register",padx=40,pady=2,fg="black", bg="grey",command=registerButton).grid(row=9, column=9)

Button(root, text="New User?",padx=30, pady=4, fg="white",bg="blue",font=("Calabri", 21),command=newUserButton).grid(row=0, column=9)

def notesSection():
    entry3=Entry(root,width=65)
    entry3.grid(row=1, column=50)
    import datetime
    current_time = datetime.datetime.now()
    entry3.insert(0, current_time)
    
def notesSection2():
    entry4=Entry(root,width=65)
    entry4.grid(row=3, column=50)
    import datetime
    current_time = datetime.datetime.now()
    entry4.insert(0, current_time)

def notesSection3():
    entry5=Entry(root,width=65)
    entry5.grid(row=5, column=50)
    import datetime
    current_time = datetime.datetime.now()
    entry5.insert(0, current_time)

def diaryMode():
    root = Tk()

    my_frame = Frame(root)
    my_frame.pack(pady=5)

    text_scroll = Scrollbar(my_frame)
    text_scroll.pack(side=RIGHT, fill=Y)

    my_text = Text(my_frame,width=97, height=25,undo=True,yscrollcommand=text_scroll.set)
    my_text.pack()

    text_scroll.config(command=my_text.yview)


root.mainloop()













